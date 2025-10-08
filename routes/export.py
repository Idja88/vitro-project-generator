import json
import os
from datetime import datetime
from io import BytesIO
from flask import Blueprint, jsonify, send_file, current_app
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import vitro_cad_api as vc
from decorators import require_token

bp = Blueprint('export', __name__, url_prefix='/export')

@bp.route('/excel/<project_id>', methods=['GET'])
@require_token
def export_project_to_excel(token, project_id):
    try:
        # Получаем данные проекта
        project = vc.get_mp_item(token, project_id)

        # Простые проверки
        if not project:
            return jsonify({'error': 'Проект не найден'}), 404
        
        if project["fieldValueMap"].get("selection_matrix") is None:
            return jsonify({'error': 'Матрица выбора отсутствует'}), 400
        
        # Путь к шаблону
        template_path = os.path.join(os.path.dirname(__file__), '..', 'templates', 'f10_template.xlsx')

        # Проверяем, существует ли файл шаблона
        if not os.path.exists(template_path):
            # Создаем Excel файл
            excel_buffer = create_simple_excel(project)
        else:
            # Если есть шаблон, используем его
            excel_buffer = create_excel_from_template(token, project, template_path)

        # Имя файла
        filename = f"{project['fieldValueMap']['name']}_Ф10.xlsx"
        
        return send_file(
            excel_buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

#Создаем файл Excel с простой структурой
def create_simple_excel(project):

    selection_matrix_raw = project["fieldValueMap"]["selection_matrix"]

    # Парсим JSON
    if isinstance(selection_matrix_raw, str):
        selection_matrix = json.loads(selection_matrix_raw)
    else:
        selection_matrix = selection_matrix_raw

    wb = Workbook()
    ws = wb.active
    ws.title = "Ф10"
    
    # Получаем объекты (не удаленные)
    objects = [obj for obj in selection_matrix['objects'] if not obj.get('deleted', False)]
    
    # Получаем все уникальные марки (учитываем ID + номер)
    all_marks = {}
    for obj in objects:
        for mark in obj['marks']:
            if not mark.get('deleted', False):
                mark_id = mark['id']
                mark_number = mark['number']
                mark_name = mark['name']
                
                # Создаем уникальный ключ: mark_id + mark_number
                mark_key = f"{mark_id}_{mark_number}"
                
                # Формируем отображаемое имя марки
                if mark_number:
                    display_name = f"{mark_name}{mark_number}"
                else:
                    display_name = mark_name
                
                all_marks[mark_key] = {
                    'id': mark_id,
                    'number': mark_number,
                    'name': mark_name,
                    'display_name': display_name
                }
    
    # Заголовки
    headers = ['Объект'] + [mark_info['display_name'] for mark_info in all_marks.values()]
    mark_keys = list(all_marks.keys())
    
    # Записываем заголовки
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Записываем данные
    for row_idx, obj in enumerate(objects, 2):
        # Имя объекта
        ws.cell(row=row_idx, column=1, value=obj['name'])
        
        # Проверяем каждую марку
        for col_idx, mark_key in enumerate(mark_keys, 2):
            mark_info = all_marks[mark_key]
            
            # Ищем точное совпадение по ID и номеру
            has_mark = any(
                mark['id'] == mark_info['id'] and 
                mark.get('number', '') == mark_info['number'] and 
                not mark.get('deleted', False)
                for mark in obj['marks']
            )
            
            ws.cell(row=row_idx, column=col_idx, value='True' if has_mark else 'False')
    
    # Сохранение
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer

# Создаем Excel файл из шаблона
def create_excel_from_template(token, project, template_path):
    
    try:
        # Загружаем шаблон
        wb = load_workbook(template_path)
        ws = wb.active
    except FileNotFoundError:
        # Fallback - создаем новый файл
        wb = Workbook()
        ws = wb.active
        ws.title = "Ф10"
    
    # Получаем данные проекта
    project_name = project["fieldValueMap"]["name"]
    project_code = project["fieldValueMap"]["code"]
    project_chief = project["fieldValueMap"]["chief_project_engineer"]["fieldValueMap"]["name"]
    selection_matrix_raw = project["fieldValueMap"]["selection_matrix"]

    # Парсим JSON
    if isinstance(selection_matrix_raw, str):
        selection_matrix = json.loads(selection_matrix_raw)
    else:
        selection_matrix = selection_matrix_raw
    
    # Получаем объекты (не удаленные)
    objects = [obj for obj in selection_matrix['objects'] if not obj.get('deleted', False)]
    
    # Получаем все уникальные марки (учитываем ID + номер)
    all_marks = {}
    for obj in objects:
        for mark in obj['marks']:
            if not mark.get('deleted', False):
                mark_id = mark['id']
                mark_number = mark['number']
                mark_name = mark['name']
                
                # Создаем уникальный ключ: mark_id + mark_number
                mark_key = f"{mark_id}_{mark_number}"
                
                # Формируем отображаемое имя марки
                if mark_number:
                    display_name = f"{mark_name}{mark_number}"
                else:
                    display_name = mark_name
                
                all_marks[mark_key] = {
                    'id': mark_id,
                    'number': mark_number,
                    'name': mark_name,
                    'display_name': display_name,
                    'full_name': None,  # Будет заполнено из API
                    'package': None  # Будет заполнено из API
                }
    
    # === ПОЛУЧЕНИЕ ПОЛНЫХ НАЗВАНИЙ МАРОК И ОТДЕЛОВ ЧЕРЕЗ API ===
    if token:
        for mark_key, mark_info in all_marks.items():
            try:
                # Получаем полную информацию о марке через API
                mark_data = vc.get_mp_item(token, mark_info['id'])
                if mark_data and 'fieldValueMap' in mark_data:
                    # Получаем название марки
                    full_name = mark_data['fieldValueMap'].get('name')
                    if full_name:
                        all_marks[mark_key]['full_name'] = full_name
                    else:
                        all_marks[mark_key]['full_name'] = mark_info['display_name']  # Fallback
                    
                    # Получаем часть марки
                    try:
                        package_name = mark_data['fieldValueMap']['sheet_package_lookup']['fieldValueMap']['name']
                        all_marks[mark_key]['package'] = package_name
                    except (KeyError, TypeError):
                        # Если части нет - оставляем None
                        all_marks[mark_key]['package'] = None

                else:
                    all_marks[mark_key]['full_name'] = mark_info['display_name']  # Fallback
                    all_marks[mark_key]['package'] = None
            except Exception as e:
                all_marks[mark_key]['full_name'] = mark_info['display_name']  # Fallback
                all_marks[mark_key]['package'] = None
    else:
        # Если нет токена, используем fallback значения
        for mark_key in all_marks:
            all_marks[mark_key]['full_name'] = all_marks[mark_key]['display_name']
            all_marks[mark_key]['package'] = None

    # === ГРУППИРОВКА МАРОК ПО ЧАСТЯМ ===
    packages_marks = {}
    marks_without_package = []

    for mark_key, mark_info in all_marks.items():
        package = mark_info['package']
        if package:
            if package not in packages_marks:
                packages_marks[package] = []
            packages_marks[package].append((mark_key, mark_info))
        else:
            # Марки без части идут в отдельный список
            marks_without_package.append((mark_key, mark_info))

    # Сортируем части по алфавиту
    sorted_packages = sorted(packages_marks.keys())

    # === ЗАПОЛНЕНИЕ МЕТАДАННЫХ ===
    # B1 - Название проекта
    ws['B1'] = project_name
    
    # F2 - Код проекта
    ws['F2'] = project_code
    
    # N3 - ГИП (главный инженер проекта)
    ws['N3'] = project_chief
    
    # S3 - Текущая дата в формате DD.MM.YYYY
    current_date = datetime.now().strftime("%d.%m.%Y")
    ws['S3'] = current_date
    
    # === РАЗМЕЩЕНИЕ ОТДЕЛОВ И МАРОК ===
    current_col = 6  # Начинаем с колонки F (6)
    mark_keys_ordered = []  # Упорядоченный список ключей марок
    
    # Сначала размещаем марки с пакетами
    for package in sorted_packages:
        marks_in_pkg = packages_marks[package]
        pkg_start_col = current_col

        # Размещаем марки пакета
        for mark_key, mark_info in marks_in_pkg:
            # Добавляем в упорядоченный список
            mark_keys_ordered.append(mark_key)
            
            # F8, G8, H8... - полные названия марок
            cell_f8 = ws.cell(row=8, column=current_col)
            cell_f8.value = mark_info['full_name']
            
            # F9, G9, H9... - коды марок
            cell_f9 = ws.cell(row=9, column=current_col)
            cell_f9.value = mark_info['display_name']
            
            # F10, G10, H10... - нумерация
            cell_f10 = ws.cell(row=10, column=current_col)
            cell_f10.value = 3 + len(mark_keys_ordered) - 1  # Нумерация с 3
            
            current_col += 1
        
        # Размещаем название части в F7 и объединяем ячейки
        pkg_end_col = current_col - 1
        if pkg_start_col <= pkg_end_col:
            # Записываем название пакета
            pkg_cell = ws.cell(row=7, column=pkg_start_col)
            pkg_cell.value = package

            # Объединяем ячейки для части (если марок больше одной)
            if pkg_start_col < pkg_end_col:
                merge_range = f"{get_column_letter(pkg_start_col)}7:{get_column_letter(pkg_end_col)}7"
                ws.merge_cells(merge_range)

    # Затем размещаем марки без части (в конце)
    for mark_key, mark_info in marks_without_package:
        # Добавляем в упорядоченный список
        mark_keys_ordered.append(mark_key)
        
        # F8, G8, H8... - полные названия марок
        cell_f8 = ws.cell(row=8, column=current_col)
        cell_f8.value = mark_info['full_name']
        
        # F9, G9, H9... - коды марок
        cell_f9 = ws.cell(row=9, column=current_col)
        cell_f9.value = mark_info['display_name']
        
        # F10, G10, H10... - нумерация
        cell_f10 = ws.cell(row=10, column=current_col)
        cell_f10.value = 3 + len(mark_keys_ordered) - 1  # Нумерация с 3
        
        current_col += 1
        
        # Для марок без отдела не заполняем строку 7 (оставляем пустой)
    
    # === РАЗМЕЩЕНИЕ ОБЪЕКТОВ ===
    for i, obj in enumerate(objects):
        row = 11 + i  # C11, C12, C13...
        
        # === ФОРМИРОВАНИЕ И РАЗМЕЩЕНИЕ КОДА ОБЪЕКТА ===
        if obj.get('number'):
            object_code = f"{project_code}-{obj['number']}"
        else:
            object_code = project_code
        
        # Записываем код объекта в столбец B
        code_cell = ws.cell(row=row, column=2)  # B=2
        code_cell.value = object_code
        
        # Записываем имя объекта в столбец C
        name_cell = ws.cell(row=row, column=3)  # C=3
        name_cell.value = obj['name']
        
        # === РАЗМЕЩЕНИЕ ПЕРЕСЕЧЕНИЙ ===
        for j, mark_key in enumerate(mark_keys_ordered):
            col = 6 + j  # F=6, G=7, H=8...
            mark_info = all_marks[mark_key]
            
            # Проверяем наличие марки у объекта
            has_mark = any(
                mark['id'] == mark_info['id'] and 
                mark.get('number', '') == mark_info['number'] and 
                not mark.get('deleted', False)
                for mark in obj['marks']
            )
            
            if has_mark:
                intersection_cell = ws.cell(row=row, column=col)
                intersection_cell.value = '+'
    
    # Сохранение
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer